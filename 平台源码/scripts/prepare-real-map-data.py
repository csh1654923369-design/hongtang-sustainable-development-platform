from __future__ import annotations

import csv
import json
import math
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
MATERIAL_ROOT = (
    Path(sys.argv[1]).resolve()
    if len(sys.argv) > 1
    else PROJECT_ROOT.parent / "平台素材"
)
OUTPUT = PROJECT_ROOT / "public" / "data" / "hongtang-real-map-features.json"
POI_WORKBOOK = MATERIAL_ROOT / "POI_1785232551999.xlsx"
POI_IMAGES = MATERIAL_ROOT / "poi图片.csv"
VILLAGE_IMAGES = MATERIAL_ROOT / "村景图片.csv"

BOUNDS = {
    "west": 99.87903589790675,
    "south": 24.626466269999977,
    "east": 99.92124683196907,
    "north": 24.685577823254807,
}


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def number(value: object) -> float:
    return float(clean(value))


def point_in_bounds(longitude: float, latitude: float) -> bool:
    return (
        BOUNDS["west"] <= longitude <= BOUNDS["east"]
        and BOUNDS["south"] <= latitude <= BOUNDS["north"]
    )


def valid_coordinate(longitude: float, latitude: float) -> bool:
    """Accept every valid WGS84 point, even when it lies outside the 2D map extent."""
    return (
        math.isfinite(longitude)
        and math.isfinite(latitude)
        and -180 <= longitude <= 180
        and -90 <= latitude <= 90
        and not (longitude == 0 and latitude == 0)
    )


def map_position(longitude: float, latitude: float) -> tuple[float, float]:
    x = (longitude - BOUNDS["west"]) / (BOUNDS["east"] - BOUNDS["west"]) * 100
    y = (BOUNDS["north"] - latitude) / (BOUNDS["north"] - BOUNDS["south"]) * 100
    return round(x, 3), round(y, 3)


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as source:
        return list(csv.DictReader(source))


def poi_feature_type(standard_code: str, standard_name: str, name: str) -> str:
    combined = f"{standard_name} {name}"
    if standard_code == "SNT0223" or "花园" in combined or "小院" in combined:
        return "garden"
    if standard_code == "SNT0901" and "茶厂" in combined:
        return "tea-factory"
    if standard_code in {"SNT0303"} or "污水" in combined or "净水" in combined:
        return "water-facility"
    if standard_code == "SNT0801" or "梯田" in combined:
        return "ecology"
    return "public-service"


def goal_for_feature(feature_type: str) -> str:
    if feature_type in {"ecology", "water-facility"}:
        return "goal-ecology"
    if feature_type == "public-service":
        return "goal-service"
    return "goal-livable"


poi_image_rows = read_csv(POI_IMAGES)
poi_images_by_name: dict[str, list[dict[str, str]]] = defaultdict(list)
for row in poi_image_rows:
    poi_images_by_name[clean(row.get("name"))].append(row)

workbook = load_workbook(POI_WORKBOOK, read_only=True, data_only=True)
worksheet = workbook.active
rows = worksheet.iter_rows(values_only=True)
headers = [clean(value) for value in next(rows)]
poi_rows = [dict(zip(headers, values)) for values in rows]
workbook.close()

poi_features: list[dict[str, object]] = []
for index, row in enumerate(poi_rows, start=1):
    longitude = number(row.get("lng"))
    latitude = number(row.get("lat"))
    if not valid_coordinate(longitude, latitude):
        continue
    name = clean(row.get("name")) or f"未命名点位{index}"
    standard_name = clean(row.get("standardName")).rstrip("；;")
    standard_code = clean(row.get("standardCode"))
    feature_type = poi_feature_type(standard_code, standard_name, name)
    images = poi_images_by_name.get(name, [])
    image_urls = list(
        dict.fromkeys(
            clean(image.get("concat('https://sannongdata.cn/AiHouseTypeData',mpv.media_url)"))
            for image in images
            if clean(image.get("concat('https://sannongdata.cn/AiHouseTypeData',mpv.media_url)"))
        )
    )
    update_times = [clean(image.get("create_time")) for image in images if clean(image.get("create_time"))]
    map_x, map_y = map_position(longitude, latitude)
    poi_features.append(
        {
            "id": f"real-poi-{index}",
            "isDemo": False,
            "title": name,
            "featureType": feature_type,
            "status": "已记录",
            "location": f"红塘村 · {standard_name or '现场点位'}",
            "description": clean(row.get("profile")) or "该点位简介待后续补充。",
            "longitude": longitude,
            "latitude": latitude,
            "mapX": map_x,
            "mapY": map_y,
            "updatedAt": max(update_times)[:10] if update_times else "资料时间未标注",
            "goalId": goal_for_feature(feature_type),
            "publicParticipation": False,
            "submittedByMe": False,
            "geometry": {"type": "Point", "coordinates": [longitude, latitude]},
            "imageLabel": f"{len(image_urls)}张现场照片" if image_urls else "现场照片待补充",
            "imageUrls": image_urls,
            "sourceLabel": "红塘村POI资料",
        }
    )

village_rows = read_csv(VILLAGE_IMAGES)
village_groups: dict[str, list[dict[str, str]]] = defaultdict(list)
for row in village_rows:
    try:
        longitude = number(row.get("lng"))
        latitude = number(row.get("lat"))
    except ValueError:
        continue
    if valid_coordinate(longitude, latitude):
        coordinate_key = f"{longitude:.8f},{latitude:.8f}"
        village_groups[coordinate_key].append(row)

village_features: list[dict[str, object]] = []
for index, (_, group) in enumerate(sorted(village_groups.items()), start=1):
    longitudes = [number(row.get("lng")) for row in group]
    latitudes = [number(row.get("lat")) for row in group]
    longitude = sum(longitudes) / len(longitudes)
    latitude = sum(latitudes) / len(latitudes)
    if not math.isfinite(longitude) or not math.isfinite(latitude):
        continue
    image_urls = list(
        dict.fromkeys(
            clean(row.get("media_url"))
            for row in group
            if clean(row.get("media_url"))
        )
    )
    update_times = [clean(row.get("create_time")) for row in group if clean(row.get("create_time"))]
    map_x, map_y = map_position(longitude, latitude)
    village_features.append(
        {
            "id": f"real-village-photo-{index}",
            "isDemo": False,
            "title": f"村景记录 {index}",
            "featureType": "research-photo",
            "status": "村景记录",
            "location": "红塘村村景记录",
            "description": f"现场村景资料，共{len(image_urls)}张照片。",
            "longitude": round(longitude, 8),
            "latitude": round(latitude, 8),
            "mapX": map_x,
            "mapY": map_y,
            "updatedAt": max(update_times)[:10] if update_times else "资料时间未标注",
            "goalId": "goal-culture",
            "publicParticipation": False,
            "submittedByMe": False,
            "geometry": {
                "type": "Point",
                "coordinates": [round(longitude, 8), round(latitude, 8)],
            },
            "imageLabel": f"{len(image_urls)}张村景照片",
            "imageUrls": image_urls,
            "sourceLabel": "红塘村村景资料",
        }
    )

payload = {
    "meta": {
        "title": "红塘村地图点位",
        "generatedAt": date.today().isoformat(),
        "source": "../平台素材/POI_1785232551999.xlsx + poi图片.csv + 村景图片.csv",
        "poiCount": len(poi_features),
        "villagePhotoPointCount": len(village_features),
        "poiImageCount": sum(len(feature["imageUrls"]) for feature in poi_features),
        "villageImageCount": sum(len(feature["imageUrls"]) for feature in village_features),
        "notice": "点位名称、简介、坐标和图片网址来自用户提供的素材；网页运行文件不包含表格之外的个人联系方式。",
    },
    "bounds": BOUNDS,
    "features": poi_features + village_features,
}

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
print(
    f"Prepared {len(poi_features)} POIs and {len(village_features)} village-photo points at {OUTPUT}"
)
